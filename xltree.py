'''XLTree output plugin.
Generates tree of YANG data and outputs to an Excel book.
'''

import optparse
import collections

from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT, PatternFill, Border, Side, Color, Alignment, Font
from openpyxl.utils import get_column_letter
from pyang import plugin
from pyang import statements


def pyang_plugin_init():
    plugin.register_plugin(XLTreePlugin())

class XLTreePlugin(plugin.PyangPlugin):
    def add_output_format(self, fmts):
        self.multiple_modules = True
        fmts['xltree'] = self

    def add_opts(self, optparser):
        optlist = [
            optparse.make_option('--xltree-out',
                                 dest='xltree_out',
                                 default='xltree.xlsx',
                                 help='Excel file to output. (default: xltree.xlsx)'),
            optparse.make_option('--xltree-font',
                                 dest='xltree_font',
                                 help='Name of the font to use. (ex. Calibri, "Yu Gothic Medium")'),
            ]

        g = optparser.add_option_group('XLTree output specific options')
        g.add_options(optlist)

    def setup_fmt(self, ctx):
        ctx.implicit_errors = False

    def emit(self, ctx, modules, fd):
        sorted_modules = sort_modules(modules)

        wb = Workbook()

        ws = wb.active
        ws.title = 'module'
        ws.sheet_view.showGridLines = False
        emit_module(ctx, sorted_modules, ws)

        ws = wb.create_sheet('data')
        ws.sheet_view.showGridLines = False
        emit_data(ctx, sorted_modules, ws)

        ws = wb.create_sheet('rpc')
        ws.sheet_view.showGridLines = False
        emit_rpc(ctx, sorted_modules, ws)

        ws = wb.create_sheet('notif')
        ws.sheet_view.showGridLines = False
        emit_notif(ctx, sorted_modules, ws)

        ws = wb.create_sheet('enum-like-type')
        ws.sheet_view.showGridLines = False
        emit_enum(ctx, sorted_modules, ws)

        ws = wb.create_sheet('identity')
        ws.sheet_view.showGridLines = False
        emit_identity(ctx, sorted_modules, ws)

        if ctx.opts.xltree_font:
            font = Font(name=ctx.opts.xltree_font)
            for ws in wb.worksheets:
                for row in ws.rows:
                    for cell in row:
                        cell.font = font

        fd.write('Output file: ')
        fd.write(ctx.opts.xltree_out)
        fd.write('\n')
        wb.save(ctx.opts.xltree_out)


MODULE_KEYS = ('name', 'stmtname', 'yang_version','namespace', 'prefix', 'revision', 'belongs_to', 'including', 'imports', 'imported_by', 'non_primary_imps', 'includes', 'included_by')
ModuleRow = collections.namedtuple('ModuleRow', MODULE_KEYS)

DATA_KEYS = ('depth', 'name', 'stmtname', 'typename', 'mandatory', 'config', 'orig_modname', 'modname', 'path_simple', 'path_keys', 'path_full', 'desc')
DataRow = collections.namedtuple('DataRow', DATA_KEYS)

ID_KEYS = ('depth', 'name', 'modname', 'fullname', 'desc')
IdRow = collections.namedtuple('IdRow', ID_KEYS)
IdRecord = collections.namedtuple('IdRecord', ['fullname', 'name', 'modname', 'basename', 'desc', 'children'])

ENUM_KEYS = ('depth', 'name', 'typename', 'modname', 'fullname', 'desc', 'children')
EnumRow = collections.namedtuple('EnumRow', ENUM_KEYS)

WIDTH_DESC = 50.0
WIDTH_LEFT_OF_NAME = 3.0

FILL_HEADER = PatternFill(patternType='solid', fgColor=Color(indexed=42))

SIDE_THIN_BLACK = Side(style='thin', color=Color(indexed=0))

HBORDER_LEFT = Border(
    left=SIDE_THIN_BLACK,
    top=SIDE_THIN_BLACK,
    bottom=SIDE_THIN_BLACK,
)

HBORDER_MID = Border(
    top=SIDE_THIN_BLACK,
    bottom=SIDE_THIN_BLACK,
)

HBORDER_RIGHT = Border(
    right=SIDE_THIN_BLACK,
    top=SIDE_THIN_BLACK,
    bottom=SIDE_THIN_BLACK,
)

HBORDER_BOX = Border(
    left=SIDE_THIN_BLACK,
    right=SIDE_THIN_BLACK,
    top=SIDE_THIN_BLACK,
    bottom=SIDE_THIN_BLACK,
)

BORDER_LEFT_OF_NAME = Border(
    left=SIDE_THIN_BLACK,
    right=SIDE_THIN_BLACK,
)

BORDER_NAME = Border(
    left=SIDE_THIN_BLACK,
    top=SIDE_THIN_BLACK,
)

BORDER_RIGHT_OF_NAME = Border(
    top=SIDE_THIN_BLACK,
)

BORDER_END_OF_NAME = Border(
    top=SIDE_THIN_BLACK,
    right=SIDE_THIN_BLACK,
)

BORDER_NAME_ADN_END = Border(
    left=SIDE_THIN_BLACK,
    right=SIDE_THIN_BLACK,
    top=SIDE_THIN_BLACK,
)

BORDER_BOX = Border(
    left=SIDE_THIN_BLACK,
    right=SIDE_THIN_BLACK,
    top=SIDE_THIN_BLACK,
    bottom=SIDE_THIN_BLACK,
)

BORDER_TOPONLY = Border(
    top=SIDE_THIN_BLACK,
)

ALGIN_TEXT = Alignment(vertical='center')
ALGIN_DESC = Alignment(vertical='center', wrap_text=True)


def emit_module(ctx, modules, ws):
    rows = gather_modules(ctx, modules)

    dim = ws.column_dimensions[get_column_letter(1)]
    dim.width = WIDTH_LEFT_OF_NAME
    dim = ws.column_dimensions[get_column_letter(2)]
    dim.width = 18.0

    non_primary_modules = set()

    nrow = 1
    for mod in rows:
        non_primary_modules.update(mod.non_primary_imps)

        ws.cell(row=nrow, column=1, value=mod.name)
        nrow += 1

        ws.cell(row=nrow, column=2, value='category:')
        ws.cell(row=nrow, column=3, value=mod.stmtname)
        nrow += 1

        ws.cell(row=nrow, column=2, value='yang-version:')
        ws.cell(row=nrow, column=3, value=mod.yang_version)
        nrow += 1

        ws.cell(row=nrow, column=2, value='namespace:')
        ws.cell(row=nrow, column=3, value=mod.namespace)
        nrow += 1

        ws.cell(row=nrow, column=2, value='prefix:')
        ws.cell(row=nrow, column=3, value=mod.prefix)
        nrow += 1

        ws.cell(row=nrow, column=2, value='revision:')
        ws.cell(row=nrow, column=3, value=mod.revision)
        nrow += 1

        ws.cell(row=nrow, column=2, value='belongs-to:')
        ws.cell(row=nrow, column=3, value=mod.belongs_to or '(none)')
        nrow += 1

        ws.cell(row=nrow, column=2, value='including-module:')
        ws.cell(row=nrow, column=3, value=mod.including or '(none)')
        nrow += 1

        ws.cell(row=nrow, column=2, value='imports:')
        if mod.imports:
            for name in mod.imports:
                ws.cell(row=nrow, column=3, value=name)
                nrow += 1
        else:
            ws.cell(row=nrow, column=3, value='(none)')
            nrow += 1

        ws.cell(row=nrow, column=2, value='imported-by:')
        if mod.imported_by:
            for name in mod.imported_by:
                ws.cell(row=nrow, column=3, value=name)
                nrow += 1
        else:
            ws.cell(row=nrow, column=3, value='(none)')
            nrow += 1

        ws.cell(row=nrow, column=2, value='includes:')
        if mod.includes:
            for name in mod.includes:
                ws.cell(row=nrow, column=3, value=name)
                nrow += 1
        else:
            ws.cell(row=nrow, column=3, value='(none)')
            nrow += 1

        ws.cell(row=nrow, column=2, value='included-by:')
        if mod.included_by:
            for name in mod.included_by:
                ws.cell(row=nrow, column=3, value=name)
                nrow += 1
        else:
            ws.cell(row=nrow, column=3, value='(none)')
            nrow += 1

    if non_primary_modules:
        ws.cell(row=nrow, column=1, value='(non-primary modules):')
        nrow += 1
        for modname in sorted(non_primary_modules):
            ws.cell(row=nrow, column=2, value=modname)
            nrow += 1


def emit_data(ctx, modules, ws):
    rows = gather_datas(modules)

    max_depth = find_max_depth(rows)

    print_dataheader(ws, max_depth)

    if rows:
        print_datarows(ws, rows, max_depth)
    else:
        print_datanone(ws, max_depth)


def emit_rpc(ctx, modules, ws):
    rows = gather_rpcs(modules)

    max_depth = find_max_depth(rows)

    print_dataheader(ws, max_depth)

    if rows:
        print_datarows(ws, rows, max_depth)
    else:
        print_datanone(ws, max_depth)


def emit_notif(ctx, modules, ws):
    rows = gather_notifs(modules)

    max_depth = find_max_depth(rows)

    print_dataheader(ws, max_depth)

    if rows:
        print_datarows(ws, rows, max_depth)
    else:
        print_datanone(ws, max_depth)


def emit_enum(ctx, modules, ws):
    idmap = gather_identities(modules)
    rows = flatten_enum(gather_enum(modules, idmap))
    idmap = None

    max_depth = find_max_depth(rows)

    print_header(ws, 'name', max_depth, 'type', 'module', 'fullname', 'description')

    if rows:
        nrow = 2
        for row in rows:
            print_enumrow(ws, nrow, row, max_depth)
            nrow += 1
        print_bottom(ws, nrow, max_depth)
    else:
        print_nodata(ws, max_depth, 'type', 'module', 'fullname', 'description')


def emit_identity(ctx, modules, ws):
    idmap = gather_identities(modules)
    rows = flatten_identity(idmap)
    idmap = None

    max_depth = find_max_depth(rows)

    print_header(ws, 'name', max_depth, 'module', 'fullname', 'description')

    if rows:
        nrow = 2
        for row in rows:
            print_idrow(ws, nrow, row, max_depth)
            nrow += 1
        print_bottom(ws, nrow, max_depth)
    else:
        print_nodata(ws, max_depth, 'module', 'fullname', 'description')


def print_header(ws, first, max_depth, *rest):
    '''First column spans to max_dpeth.
    Last column is for description.
    '''

    ncol = 1
    cell = ws.cell(row=1, column=ncol, value=first)
    cell.border = HBORDER_LEFT
    cell.fill = FILL_HEADER

    for ncol in range(2, max_depth):
        cell = ws.cell(row=1, column=ncol)
        cell.border = HBORDER_MID
        cell.fill = FILL_HEADER

    ncol = max_depth
    cell = ws.cell(row=1, column=ncol)
    cell.border = HBORDER_RIGHT
    cell.fill = FILL_HEADER

    for title in rest:
        ncol = ncol + 1
        cell = ws.cell(row=1, column=ncol, value=title)
        cell.border = HBORDER_BOX
        cell.fill = FILL_HEADER

    for ncol in range(1, max_depth):
        dim = ws.column_dimensions[get_column_letter(ncol)]
        dim.width = WIDTH_LEFT_OF_NAME
    dim = ws.column_dimensions[get_column_letter(max_depth + len(rest))]
    dim.width = WIDTH_DESC


def print_nodata(ws, max_depth, *rest):
    ncol = 1
    cell = ws.cell(row=2, column=ncol, value='(no data)')
    cell.border = HBORDER_LEFT

    for ncol in range(2, max_depth):
        cell = ws.cell(row=2, column=ncol)
        cell.border = HBORDER_MID

    ncol = max_depth
    cell = ws.cell(row=2, column=ncol)
    cell.border = HBORDER_RIGHT

    for _ in rest:
        ncol = ncol + 1
        cell = ws.cell(row=2, column=ncol)
        cell.border = HBORDER_BOX


def print_name(ws, nrow, row, max_depth):
    '''Print first column to max_depth,
    with value in appropriate depth.
    '''

    ncol = 1
    for _ in range(ncol, row.depth):
        cell = ws.cell(row=nrow, column=ncol)
        cell.border = BORDER_LEFT_OF_NAME
        ncol += 1

    cell = ws.cell(row=nrow, column=ncol, value=row.name)
    cell.alignment = ALGIN_TEXT
    if ncol == max_depth:
        cell.border = BORDER_NAME_ADN_END
        ncol += 1
    else:
        cell.border = BORDER_NAME
        ncol += 1

        for _ in range(ncol, max_depth):
            cell = ws.cell(row=nrow, column=ncol)
            cell.border = BORDER_RIGHT_OF_NAME
            ncol += 1

        # should be ncol == max_depth
        cell = ws.cell(row=nrow, column=ncol)
        cell.border = BORDER_END_OF_NAME
        ncol += 1


def print_text(ws, nrow, ncol, data):
    '''Print a text to the cell.
    '''

    cell = ws.cell(row=nrow, column=ncol, value=data)
    cell.border = BORDER_BOX
    cell.alignment = ALGIN_TEXT


def print_desc(ws, nrow, ncol, data):
    '''Print description with word-wrap.
    '''

    cell = ws.cell(row=nrow, column=ncol, value=xls_escape_desc(data))
    cell.border = BORDER_BOX
    cell.alignment = ALGIN_DESC


def print_dataheader(ws, max_depth):
    # M/C/O/c : Mandatory, Mandatory with when, Optional, Optional with when
    print_header(ws, 'name', max_depth, 'keyword', 'type', 'M/C/O/c', 'mode', 'orig-module', 'module', 'path(simple)', 'path(keys)', 'path(full)', 'description')


def print_datanone(ws, max_depth):
    print_nodata(ws, max_depth, 'keyword', 'type', 'M/C/O/c', 'mode', 'orig-module', 'module', 'path(simple)', 'path(keys)', 'path(full)', 'description')


def print_datarows(ws, rows, max_depth):

    nrow = 2
    for row in rows:
        print_name(ws, nrow, row, max_depth)
        ncol = max_depth + 1

        print_text(ws, nrow, ncol, row.stmtname)
        ncol += 1

        print_text(ws, nrow, ncol, row.typename)
        ncol += 1

        print_text(ws, nrow, ncol, row.mandatory)
        ncol += 1

        print_text(ws, nrow, ncol, row.config)
        ncol += 1

        print_text(ws, nrow, ncol, row.orig_modname)
        ncol += 1

        print_text(ws, nrow, ncol, row.modname)
        ncol += 1

        print_text(ws, nrow, ncol, row.path_simple)
        ncol += 1

        print_text(ws, nrow, ncol, row.path_keys)
        ncol += 1

        print_text(ws, nrow, ncol, row.path_full)
        ncol += 1

        print_desc(ws, nrow, ncol, row.desc)

        nrow += 1

    print_bottom(ws, nrow, max_depth)


def print_enumrow(ws, nrow, row, max_depth):

    print_name(ws, nrow, row, max_depth)
    ncol = max_depth + 1

    print_text(ws, nrow, ncol, row.typename)
    ncol += 1

    print_text(ws, nrow, ncol, row.modname)
    ncol += 1

    print_text(ws, nrow, ncol, row.fullname)
    ncol += 1

    print_desc(ws, nrow, ncol, row.desc)


def print_idrow(ws, nrow, row, max_depth):

    print_name(ws, nrow, row, max_depth)
    ncol = max_depth + 1

    print_text(ws, nrow, ncol, row.modname)
    ncol += 1

    print_text(ws, nrow, ncol, row.fullname)
    ncol += 1

    print_desc(ws, nrow, ncol, row.desc)


def print_bottom(ws, nrow, max_depth):
    for ncol in range(1, max_depth + 1):
        cell = ws.cell(row=nrow, column=ncol)
        cell.border = BORDER_TOPONLY


def sort_modules(modules):
    return sorted(modules, key=lambda e: e.i_modulename)


def find_max_depth(rows):
    if len(rows) == 0:
        return 1
    else:
        max_depth = 0
        for row in rows:
            max_depth = max(max_depth, row.depth)
        return max_depth


def escape_desc(text):
    if text is None:
        return '-'
    else:
        return '"' + text.replace('\u2013', '-') + '"'


def xls_escape_desc(text):
    if text is None:
        return '-'
    else:
        return text.replace('\u2013', '-')


def get_root_type(stmt):
    type_obj = stmt.search_one('type')
    typedef_obj = getattr(type_obj, 'i_typedef', None)
    if typedef_obj:
        type_obj = get_root_type(typedef_obj)
    return type_obj


def is_mandatory_node(stmt):
    '''Different from statements.py that mandaotry is considered for non-config data also.
    Operational(config=false) data as well as config data should have
    mandatory/optional to indicate whether server should always supply the value or not.
    '''
    kw = stmt.keyword
    if kw in ('leaf', 'choice', 'anyxml', 'anydata'):
        # mandatory if mandatory is true
        elem = stmt.search_one('mandatory')
        if elem is not None and elem.arg == 'true':
            return True
    elif kw in ('list', 'leaf-list'):
        # mandatory if min #elements is gte 1
        elem = stmt.search_one('min-elements')
        if elem is not None and elem.arg.isnumeric() and int(elem.arg) > 0:
            return True
    elif kw == 'container':
        # mandatory if it's not a presence container
        # and at least one of the descendants is mandatory
        if stmt.search_one('presence') is None:
            for elem in stmt.i_children:
                if is_mandatory_node(elem):
                    return True
    return False


def has_when(stmt):
    '''Returns true if the statement or the source augment statement has when substatement.'''
    if stmt.search_one('when'):
        return True
    elif hasattr(stmt, 'i_augment') and stmt.i_augment.search_one('when'):
        return True
    return False


def get_mandatory_value(stmt):
    '''Returns value for the mandatory field.'''
    mandatory = is_mandatory_node(stmt)
    if has_when(stmt):
        if mandatory:
            return 'C'
        else:
            return 'c'
    else:
        if mandatory:
            return 'M'
        else:
            return 'O'


def path_for_debug(stmt):
    if stmt.keyword in ('module', 'submodule'):
        return stmt.i_modulename
    else:
        statements.mk_path_str(stmt, prefix_onchange=True, prefix_to_module=True)


def find_unresolved_uses_and_exit(node):
    for elem in node.search('uses'):
        if elem.i_grouping is None:
            raise RuntimeError('unresolved uses from {} to {}'.format(path_for_debug(node), elem.arg))


def gather_modules(ctx, modules):
    modmap = {}
    for module in modules:
        mod = ModuleRow(
            module.i_modulename,
            module.keyword,
            module.i_version,
            getattr(module.search_one('namespace'), 'arg', ''),
            getattr(module.search_one('prefix'), 'arg', ''),
            module.i_latest_revision,
            getattr(module.search_one('belongs-to'), 'arg', ''),
            module.i_including_modulename or '',
            [], [], [], [], [])
        for stmt in module.search('import'):
            impname = stmt.arg
            impmod = ctx.get_module(impname)
            if not impmod:
                raise RuntimeError('imported {} was not found'.format(impname))
            mod.imports.append(impname)
            if not impmod.i_is_primary_module:
                mod.non_primary_imps.append(impname)
        for stmt in module.search('include'):
            subname = stmt.arg
            if not ctx.get_module(subname):
                raise RuntimeError('included {} was not found'.format(subname))
            mod.includes.append(subname)
        modmap[mod.name] = mod

    for mod in modmap.values():
        for impname in mod.imports:
            # impname may be i_is_primary_module == False
            if impname in modmap:
                modmap[impname].imported_by.append(mod.name)
        for subname in mod.includes:
            # submodule should exist
            modmap[subname].included_by.append(mod.name)

    return sorted(modmap.values(), key=lambda e: e.name)


def gather_datas(modules):
    rows = []
    for module in modules:

        for elem in module.search('augment'):
            if elem.i_target_node is None:
                raise RuntimeError('unresolved autgment from {} to {}'.format(module.i_modulename, elem.arg))

        find_unresolved_uses_and_exit(module)
        chs = [ch for ch in module.i_children
               if ch.keyword in statements.data_definition_keywords]
        # data_definition_keywords =
        # ['container', 'leaf', 'leaf-list', 'list', 'case',
        #  'choice', 'anyxml', 'anydata', 'uses', 'augment']
        gather_children(rows, chs, 'data', 1)

    return rows


def gather_rpcs(modules):
    rows = []
    for module in modules:
        gather_children(rows, module.search('rpc'), 'rpc', 1)

    return rows


def gather_notifs(modules):
    rows = []
    for module in modules:
        gather_children(rows, module.search('notification'), 'notification', 1)

    return rows


def gather_children(rows, i_children, mode, depth):
    for node in i_children:
        gather_node(rows, node, mode, depth)


def gather_node(rows, node, mode, depth):

    find_unresolved_uses_and_exit(node)

    if mode == 'rpc':
        if node.keyword == 'input':
            mode = 'input'
        elif node.keyword == 'output':
            mode = 'output'
    elif node.keyword == 'action':
        mode = 'rpc'

    rows.append(create_datarow(depth, node, mode))

    gather_children(rows, getattr(node, 'i_children', []), mode, depth + 1)


def create_datarow(depth, stmt, mode):

    stmtname = stmt.keyword
    desc = statements.get_description(stmt)

    typename = '-'
    # type_desc = ''

    typestmt = stmt.search_one('type')
    if typestmt is not None:
        # type_desc = statements.get_description(typestmt) or ''
        if typestmt.i_typedef is None and typestmt.i_type_spec is None:
            raise RuntimeError('unresolved type {} at {}'.format(typestmt.arg, path_for_debug(stmt)))
        typename = statements.get_qualified_type(stmt)

        if typename == 'enumeration':
            typename = '{} : {{{}}}'.format(typename, ','.join([elem.arg for elem in typestmt.substmts]))
        elif typename == 'leafref':
            elem = typestmt.search_one('path')
            if elem is not None:
                typename = '{} : {}'.format(typename, elem.arg)
        elif typename == 'identityref':
            elem = typestmt.search_one('base')
            if elem is not None:
                typename = '{} {{{}}}'.format(typename, elem.arg)

    if stmtname in ('leaf', 'choice', 'anyxml', 'anydata'):
        mandatory = get_mandatory_value(stmt)
    elif stmtname in ('list', 'leaf-list', 'container'):
        mandatory = get_mandatory_value(stmt)
        mandatory = '-' if mandatory == 'O' else mandatory
    else:
        # case, notification, rpc, input, output, etc.,
        mandatory = '-'

    if mode == 'input':
        config = 'w'
    elif stmtname in ('rpc', 'action'):
        config = 'x'
    elif stmtname == 'notification':
        config = 'n'
    elif not stmt.i_config or mode in ('output', 'notification'):
        config = 'ro'
    else:
        config = 'rw'

    name = stmt.arg
    if stmtname == 'choice':
        name = '({})'.format(name)
    elif stmtname == 'case':
        name = ':({})'.format(name)
    elif stmtname == 'container':
        if stmt.search_one('presence') is not None:
            stmtname = '{}(p)'.format(stmtname)
            name = name + '!'
    elif stmtname in ('list', 'leaf-list'):
        if stmtname == 'list':
            keys = statements.get_keys(stmt)
            name = '{}[{}]'.format(name, ' '.join(keys))
        else:
            name = name + '*'

        elem = stmt.search_one('min-elements')
        if elem is not None:
            min = elem.arg
        else:
            min = None
        elem = stmt.search_one('max-elements')
        if elem is not None:
            max = elem.arg
        else:
            max = None

        if min is None and max is None:
            pass
        elif min is None:
            stmtname = '{}[{}..{}]'.format(stmtname, 0, max)
        elif max is None:
            stmtname = '{}[{}..{}]'.format(stmtname, min, '*')
        elif min == max:
            stmtname = '{}[{}]'.format(stmtname, min)
        else:
            stmtname = '{}[{}..{}]'.format(stmtname, min, max)

    if hasattr(stmt, 'i_orig_module'):
        orig_modname = stmt.i_orig_module.i_modulename
    else:
        # input and output stmt *sometimes* does not have i_orig_module
        orig_modname = stmt.i_module.i_modulename

    path_simple = statements.mk_path_str(stmt)
    path_keys = statements.mk_path_str(
            stmt,
            with_prefixes=False,
            prefix_onchange=False,
            prefix_to_module=True,
            resolve_top_prefix_to_module=False,
            with_keys=True)
    path_full = statements.mk_path_str(
            stmt,
            with_prefixes=False,
            prefix_onchange=True,
            prefix_to_module=True,
            resolve_top_prefix_to_module=False,
            with_keys=True)

    return DataRow(
        depth = depth,
        name = name,
        stmtname = stmtname,
        typename = typename,
        mandatory =  mandatory,
        config = config,
        orig_modname = orig_modname,
        modname = stmt.i_module.i_modulename,
        path_simple = path_simple,
        path_keys = path_keys,
        path_full = path_full,
        desc = desc,
    )


def gather_enum(modules, idmap):
    rows = []
    for module in modules:
        modname = module.i_modulename
        for stmt in module.search('typedef'):
            typeobj = get_root_type(stmt)
            match getattr(typeobj, 'arg', None):
                case 'enumeration':
                    children = []
                    for child in typeobj.search('enum'):
                        children.append(EnumRow(2, child.arg, child.keyword, child.i_module.i_modulename, '-', statements.get_description(child), None))
                    rows.append(EnumRow(1, stmt.arg, typeobj.arg, modname, ':'.join((modname, stmt.arg)), statements.get_description(stmt), children))
                case 'identityref':
                    children = []
                    base = typeobj.search_one('base')
                    identity = idmap[':'.join((base.i_module.i_modulename, base.arg))]
                    def flatten(identity):
                        children.append(EnumRow(2, identity.name, 'identity', identity.modname, identity.fullname, identity.desc, None))
                        for child in identity.children:
                            flatten(child)
                    flatten(identity)
                    rows.append(EnumRow(1, stmt.arg, typeobj.arg, modname, ':'.join((modname, stmt.arg)), statements.get_description(stmt), sorted(children, key=lambda e: e.name)))

    return rows


def flatten_enum(rows):
    sorted_rows = []
    for enum in sorted(rows, key=lambda e: e.name):
        sorted_rows.append(enum)
        for child in enum.children:
            sorted_rows.append(child)

    return sorted_rows


def gather_identities(modules):
    idmap = {}
    for module in modules:
        modname = module.i_modulename
        for stmt in module.search('identity'):
            basename = None
            base = stmt.search_one('base')
            if base:
                base = base.i_identity
                basename = ':'.join((base.i_module.i_modulename, base.arg))

            fullname = ':'.join((modname, stmt.arg))
            idmap[fullname] = IdRecord(fullname, stmt.arg, modname, basename, statements.get_description(stmt), [])

    for idrec in idmap.values():
        if idrec.basename:
            idmap[idrec.basename].children.append(idrec)

    return idmap


def flatten_identity(idmap):
    bases = []
    for idrec in idmap.values():
        if not idrec.basename:
            bases.append(idrec)

    rows = []
    def flatten(curlist, depth):
        for val in sorted(curlist, key=lambda e: e.name):
            rows.append(IdRow(depth, val.name, val.modname, val.fullname, val.desc))
            flatten(val.children, depth + 1)

    flatten(bases, 1)
    return rows
